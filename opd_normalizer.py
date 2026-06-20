"""
OPD Excel Normalizer
====================
Bhagirathi Health Care - Diagnostic Lab Management Tool

Normalizes Sales Report + Cash Register XLS exports into a clean,
matched Excel workbook with full traceability.

Author  : Built for Bhagirathi Health Care
Requires: Python 3.12, pandas, openpyxl, xlrd, tkinter
"""

import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import datetime
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─────────────────────────── helpers ────────────────────────────────────────

def get_financial_year(dt: datetime.datetime) -> str:
    """
    Return 4-char FY code for a given datetime.
    Apr 2025–Mar 2026 → '2526'
    Apr 2026–Mar 2027 → '2627'
    """
    if dt.month >= 4:
        y1, y2 = dt.year, dt.year + 1
    else:
        y1, y2 = dt.year - 1, dt.year
    return f"{str(y1)[-2:]}{str(y2)[-2:]}"


def extract_date_from_remarks(text: str) -> datetime.datetime | None:
    """
    Try to find a date inside a Remarks string.
    Handles patterns like: 'Due Amt For 15/03/2026', 'Old due for 22/05/2026'
    Returns a datetime or None.
    """
    if not text:
        return None
    # Match dd/mm/yyyy
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', str(text))
    if m:
        try:
            return datetime.datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def parse_ddmmyyyy(text) -> datetime.datetime | None:
    """
    Parse a 'dd/mm/yyyy' style string (the date-block header that appears
    once at the top of each day's group in Column A of the Cash Register)
    into a datetime. Returns None if it doesn't match.
    """
    s = safe_str(text)
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if not m:
        return None
    try:
        return datetime.datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def is_case_number(val) -> bool:
    """Return True if val looks like 2627/00594 (FY/NNNNN)."""
    return bool(re.fullmatch(r'\d{4}/\d{5}', str(val).strip()))


def is_date_like(val) -> bool:
    """Return True if val is a datetime object or a date-string."""
    return isinstance(val, datetime.datetime)


def safe_str(val) -> str:
    return "" if (val is None or (isinstance(val, float) and str(val) == "nan")) else str(val).strip()


# ─────────────────────────── parsers ────────────────────────────────────────

def parse_sales_report(path: str) -> list[dict]:
    """
    Parse the hierarchical Sales Report XLS.

    Structure per case (all in Column A, amounts in Column F):
        Row N   : datetime  → Case Date
        Row N+1 : YYYY/NNNNN → Case Number
        Row N+2 : <name>    → Patient Name
        Row N+k : <test>    → Test Name  (Col F = amount)
        ... more tests until next datetime row ...

    Returns a list of dicts:
        {case_date, case_number, patient_name, test_name, test_amount}
    """
    engine = "xlrd" if path.lower().endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, engine=engine, header=None)

    records = []
    current_date = None
    current_case = None
    current_patient = None
    reading_tests = False

    for _, row in df.iterrows():
        col_a = row.iloc[0] if len(row) > 0 else None
        col_f = row.iloc[5] if len(row) > 5 else None

        a = safe_str(col_a)

        # Skip fully blank rows
        if not a:
            reading_tests = False
            continue

        # Detect a date row (datetime object in col A)
        if is_date_like(col_a):
            current_date = col_a
            current_case = None
            current_patient = None
            reading_tests = False
            continue

        # Detect a case-number row — patient name is in col 2 on the same row
        if is_case_number(a):
            current_case = a
            # Patient name is stored in column C (index 2) on the same row
            col_c = row.iloc[2] if len(row) > 2 else None
            current_patient = safe_str(col_c) if col_c and safe_str(col_c) else None
            reading_tests = (current_patient is not None)
            continue

        # Fallback: if patient name was not captured yet, first non-blank becomes name
        if current_date and current_case and current_patient is None:
            current_patient = a
            reading_tests = True
            continue

        # Collect test rows
        if reading_tests and current_date and current_case and current_patient:
            try:
                amount = float(col_f) if col_f not in (None, "") and safe_str(col_f) else 0.0
            except (ValueError, TypeError):
                amount = 0.0

            records.append({
                "case_date": current_date,
                "case_number": current_case,
                "patient_name": current_patient,
                "test_name": a,
                "test_amount": amount,
            })

    return records


def parse_cash_register(path: str) -> list[dict]:
    """
    Parse the Cash Register XLS.

    IMPORTANT — date layout:
        Column A is NOT a per-row transaction date. It only holds a date
        string (e.g. '17/06/2026') on the single header row that starts
        each day's block of entries. Every row under that header has a
        plain serial number in Column A instead, right up until the next
        day's date header appears. THAT Column A date is the real
        transaction (cash receipt) date and must be forward-filled down
        onto every row in its block.

        Column B ('Case Dt') is a different thing — it's the date of the
        underlying case/test, which can be earlier than the receipt date
        (e.g. an old due collected today). It's kept separately as
        'case_date' rather than used as the transaction date.

    Data rows have:
        Col 0  : serial no  (int)              — NOT the date
        Col 1  : case date  (datetime)          ← "Case Dt" in header, original case date
        Col 2  : trans no   (e.g. M-000881/26-27)
        Col 3  : patient name
        Col 5  : case number (e.g. 00594)
        Col 7  : income
        Col 8  : entry_by / discount label   ← careful: col 8 = Entry By in header
        Col 9  : expense / income amount
        Col 10 : discount amount
        Col 12 : remarks (free text)

    Observed real column mapping from header row:
        0=Sl No, 1=Case Dt, 2=Trans No, 3=Name/Particular,
        4=Case No, 5=Bill Amt, 6=Entry By, 7=Income, 8=Disc, 9=Expense, 10=Remarks

    But actual data shows Case No in col 5, Entry By in col 8, Income col 7,
    Discount in col 10, Remarks in col 12. We'll use dynamic header detection.
    """
    engine = "xlrd" if path.lower().endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, engine=engine, header=None)

    records = []
    current_trans_date = None  # forward-filled from the Column A date-block header

    for _, row in df.iterrows():
        col0 = row.iloc[0] if len(row) > 0 else None
        col1 = row.iloc[1] if len(row) > 1 else None  # case date (Case Dt)
        col2 = row.iloc[2] if len(row) > 2 else None  # trans no
        col3 = row.iloc[3] if len(row) > 3 else None  # patient name
        col5 = row.iloc[5] if len(row) > 5 else None  # case no (raw)
        col7 = row.iloc[7] if len(row) > 7 else None  # bill amount (ignored)
        col8 = row.iloc[8] if len(row) > 8 else None  # entry_by
        col9 = row.iloc[9] if len(row) > 9 else None  # income (actual received — col J in source)
        col10 = row.iloc[10] if len(row) > 10 else None  # discount
        col12 = row.iloc[12] if len(row) > 12 else None  # remarks

        # A new date-block header row looks like '17/06/2026' as a plain
        # string in Column A. Capture it and move to the next row — it
        # carries no transaction data of its own.
        block_date = parse_ddmmyyyy(col0)
        if block_date is not None:
            current_trans_date = block_date
            continue

        # A data row has a numeric serial number in col 0 (not a date)
        if col0 is None or not isinstance(col0, (int, float)):
            continue
        try:
            serial = int(col0)
        except (ValueError, TypeError):
            continue
        if serial <= 0:
            continue

        # Skip data rows we haven't seen a date-block header for yet
        if current_trans_date is None:
            continue
        trans_date = current_trans_date

        # Case date — the per-row 'Case Dt' value (Column B), kept separate
        # from the transaction date.
        case_date = col1 if is_date_like(col1) else None

        # Case number: should be a 5-digit string like '00594'
        raw_case = safe_str(col5)
        if not re.fullmatch(r'\d{5}', raw_case):
            continue

        # Remarks: col 12 (free text)
        remarks = safe_str(col12)

        # Determine FY from remarks date if present, else from the case
        # date (cases are numbered by the FY they were opened in), falling
        # back to the transaction date if no case date is available.
        remarks_date = extract_date_from_remarks(remarks)
        fy_date = remarks_date or case_date or trans_date
        fy = get_financial_year(fy_date)

        full_case_number = f"{fy}/{raw_case}"

        # Transaction number: strip the /YY-YY suffix if present
        trans_no = safe_str(col2).split("/")[0] if col2 else ""
        # Keep full format e.g. M-000881 (drop /26-27)
        trans_no_clean = re.sub(r'/\d{2}-\d{2}$', '', safe_str(col2))

        try:
            income = float(col9) if col9 not in (None, "") and safe_str(col9) else 0.0
        except (ValueError, TypeError):
            income = 0.0

        try:
            discount = float(col10) if col10 not in (None, "") and safe_str(col10) else 0.0
        except (ValueError, TypeError):
            discount = 0.0

        entry_by = safe_str(col8)
        patient_name = safe_str(col3)

        records.append({
            "trans_date": trans_date,
            "case_date": case_date,
            "trans_number": trans_no_clean,
            "patient_name": patient_name,
            "case_number": full_case_number,
            "entry_by": entry_by,
            "income": income,
            "discount": discount,
            "remarks": remarks,
        })

    return records


# ─────────────────────────── export ─────────────────────────────────────────

def auto_width(ws):
    """Auto-fit all column widths based on cell content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def style_header_row(ws, row_num: int, bg_color: str = "1F4E79"):
    """Apply header styling to a specific row."""
    hdr_fill = PatternFill("solid", fgColor=bg_color)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[row_num]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border


def export_to_excel(
    sales_records: list[dict],
    cash_records: list[dict],
    output_path: str,
):
    """
    Write two-sheet Excel output.

    Sheet 1 — sales_report  : one row per test
        Columns: Case Date | Case Number | Patient Name | Test Name | Test Amount

    Sheet 2 — cash_register : one row per cash entry (Bill Amount excluded)
        Columns: Transaction Date | Case Date | Transaction No | Case No |
                 Patient Name | Entry By | Income | Discount
    """
    wb = Workbook()

    norm_font  = Font(size=10)
    center_al  = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="center")

    # ── Sheet 1: sales_report ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "sales_report"
    ws1.freeze_panes = "A2"

    ws1.append(["Case Date", "Case Number", "Patient Name", "Test Name", "Test Amount"])
    style_header_row(ws1, 1, "1F4E79")   # dark navy header

    alt_blue   = PatternFill("solid", fgColor="EBF3FB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    for i, r in enumerate(sales_records, start=2):
        case_dt = (r["case_date"].strftime("%d-%m-%Y")
                   if isinstance(r["case_date"], datetime.datetime)
                   else safe_str(r["case_date"]))
        ws1.append([
            case_dt,
            r["case_number"],
            r["patient_name"],
            r["test_name"],
            r["test_amount"],
        ])
        fill = alt_blue if i % 2 == 0 else white_fill
        for col_idx, cell in enumerate(ws1[i], start=1):
            cell.fill = fill
            cell.font = norm_font
            # Centre date, case no, amount; left-align names
            cell.alignment = center_al if col_idx in (1, 2, 5) else left_al

    auto_width(ws1)

    # ── Sheet 2: cash_register ────────────────────────────────────────────
    ws2 = wb.create_sheet("cash_register")
    ws2.freeze_panes = "A2"

    ws2.append([
        "Transaction Date", "Case Date", "Transaction No", "Case No",
        "Patient Name", "Entry By", "Income", "Discount",
    ])
    style_header_row(ws2, 1, "174E3B")   # dark green header

    alt_green  = PatternFill("solid", fgColor="EBF7F0")
    cream_fill = PatternFill("solid", fgColor="FFFFFF")

    for i, r in enumerate(cash_records, start=2):
        trans_dt = (r["trans_date"].strftime("%d-%m-%Y")
                    if isinstance(r["trans_date"], datetime.datetime)
                    else safe_str(r["trans_date"]))
        case_dt = (r["case_date"].strftime("%d-%m-%Y")
                   if isinstance(r["case_date"], datetime.datetime)
                   else safe_str(r["case_date"]))
        ws2.append([
            trans_dt,
            case_dt,
            r["trans_number"],
            r["case_number"],
            r["patient_name"],
            r["entry_by"],
            r["income"],
            r["discount"],
        ])
        fill = alt_green if i % 2 == 0 else cream_fill
        for col_idx, cell in enumerate(ws2[i], start=1):
            cell.fill = fill
            cell.font = norm_font
            cell.alignment = center_al if col_idx in (1, 2, 3, 4, 7, 8) else left_al

    auto_width(ws2)
    wb.save(output_path)


# ─────────────────────────── normalizer core ────────────────────────────────

def normalize(sales_path: str, cash_path: str, output_path: str, log_fn=None):
    """
    Main normalization pipeline.
    log_fn(msg) → optional progress callback.
    Returns (sales_records, cash_records).
    """

    def log(msg):
        if log_fn:
            log_fn(msg)

    log("📂 Reading Sales Report…")
    sales_records = parse_sales_report(sales_path)
    log(f"   → {len(sales_records)} test rows")
    log(f"   → {len(set(r['case_number'] for r in sales_records))} unique cases")

    log("📂 Reading Cash Register…")
    cash_records = parse_cash_register(cash_path)
    log(f"   → {len(cash_records)} cash entries")

    log("💾 Writing Excel file…")
    export_to_excel(sales_records, cash_records, output_path)
    log(f"✅ Done!  Sales rows: {len(sales_records)}  |  Cash rows: {len(cash_records)}")
    log(f"   Saved to: {output_path}")
    return sales_records, cash_records


# ─────────────────────────── GUI ────────────────────────────────────────────

class OPDNormalizerApp(tk.Tk):

    BG = "#0F172A"          # dark navy
    CARD = "#1E293B"        # card bg
    ACCENT = "#38BDF8"      # sky blue
    ACCENT2 = "#0EA5E9"     # deeper blue
    TEXT = "#F1F5F9"        # near-white text
    MUTED = "#94A3B8"       # muted text
    SUCCESS = "#22C55E"     # green
    ERROR = "#EF4444"       # red
    WARN = "#F59E0B"        # amber
    BORDER = "#334155"      # border colour

    def __init__(self):
        super().__init__()
        self.title("OPD Excel Normalizer — Bhagirathi Health Care")
        self.geometry("780x680")
        self.resizable(True, True)
        self.configure(bg=self.BG)
        self.minsize(700, 580)

        self.sales_path = tk.StringVar()
        self.cash_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self._build_ui()
        self._set_default_output()

    # ── UI construction ───────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header ──
        hdr = tk.Frame(self, bg="#0EA5E9", pady=14)
        hdr.pack(fill="x")
        tk.Label(
            hdr, text="🏥  OPD Excel Normalizer",
            font=("Segoe UI", 18, "bold"),
            bg="#0EA5E9", fg="white",
        ).pack()
        tk.Label(
            hdr, text="Bhagirathi Health Care — Diagnostic Lab",
            font=("Segoe UI", 10), bg="#0EA5E9", fg="#E0F2FE",
        ).pack()

        # ── Main card ──
        card = tk.Frame(self, bg=self.CARD, bd=0, padx=28, pady=22)
        card.pack(fill="both", expand=True, padx=20, pady=18)

        # File selectors
        self._file_row(card, "Sales Report (.xls / .xlsx):", self.sales_path,
                       self._browse_sales, row=0)
        self._file_row(card, "Cash Register (.xls / .xlsx):", self.cash_path,
                       self._browse_cash, row=1)
        self._file_row(card, "Output file (.xlsx):", self.output_path,
                       self._browse_output, row=2, is_save=True)

        # Separator
        sep = tk.Frame(card, bg=self.BORDER, height=1)
        sep.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(18, 14))

        # Normalize button
        self.btn_run = tk.Button(
            card, text="⚡  Normalize & Export",
            font=("Segoe UI", 12, "bold"),
            bg=self.ACCENT2, fg="white",
            activebackground="#0284C7", activeforeground="white",
            relief="flat", padx=20, pady=10,
            cursor="hand2",
            command=self._run,
        )
        self.btn_run.grid(row=4, column=0, columnspan=3, pady=(0, 14), sticky="ew")

        # Progress bar
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "BNH.Horizontal.TProgressbar",
            troughcolor=self.BG, background=self.ACCENT2,
            thickness=8,
        )
        self.progress = ttk.Progressbar(
            card, mode="indeterminate", style="BNH.Horizontal.TProgressbar",
        )
        self.progress.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 12))

        # Log area label
        tk.Label(
            card, text="Processing Log",
            font=("Segoe UI", 10, "bold"),
            bg=self.CARD, fg=self.MUTED,
        ).grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 4))

        # Log text area
        log_frame = tk.Frame(card, bg=self.BORDER, padx=1, pady=1)
        log_frame.grid(row=7, column=0, columnspan=3, sticky="nsew")
        card.rowconfigure(7, weight=1)

        self.log_text = tk.Text(
            log_frame,
            font=("Consolas", 9),
            bg="#0D1B2A", fg="#7DD3FC",
            insertbackground=self.ACCENT,
            relief="flat",
            wrap="word",
            state="disabled",
            padx=10, pady=8,
        )
        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview, bg=self.BORDER)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

        # Configure tag colours for log
        self.log_text.tag_config("success", foreground="#22C55E")
        self.log_text.tag_config("error", foreground="#EF4444")
        self.log_text.tag_config("warn", foreground="#F59E0B")
        self.log_text.tag_config("info", foreground="#7DD3FC")

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = tk.Label(
            self, textvariable=self.status_var,
            font=("Segoe UI", 9),
            bg=self.BORDER, fg=self.MUTED,
            anchor="w", padx=12,
        )
        status_bar.pack(fill="x", side="bottom")

        card.columnconfigure(1, weight=1)

    def _file_row(self, parent, label, var, command, row, is_save=False):
        tk.Label(
            parent, text=label,
            font=("Segoe UI", 10, "bold"),
            bg=self.CARD, fg=self.TEXT,
            anchor="w",
        ).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=6)

        entry = tk.Entry(
            parent, textvariable=var,
            font=("Segoe UI", 9),
            bg="#0F172A", fg=self.ACCENT,
            insertbackground=self.ACCENT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=self.BORDER,
            highlightcolor=self.ACCENT,
        )
        entry.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=6, ipady=6)

        btn = tk.Button(
            parent, text="Browse…",
            font=("Segoe UI", 9),
            bg=self.BORDER, fg=self.TEXT,
            activebackground=self.ACCENT2, activeforeground="white",
            relief="flat", padx=10, pady=4,
            cursor="hand2",
            command=command,
        )
        btn.grid(row=row, column=2, pady=6, sticky="ew")

    # ── File dialogs ──────────────────────────────────────────────────────

    def _browse_sales(self):
        p = filedialog.askopenfilename(
            title="Select Sales Report",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if p:
            self.sales_path.set(p)
            self._set_default_output()

    def _browse_cash(self):
        p = filedialog.askopenfilename(
            title="Select Cash Register",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
        )
        if p:
            self.cash_path.set(p)

    def _browse_output(self):
        p = filedialog.asksaveasfilename(
            title="Save Normalized Output As",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="normalized_output.xlsx",
        )
        if p:
            self.output_path.set(p)

    def _set_default_output(self):
        if self.sales_path.get():
            folder = os.path.dirname(self.sales_path.get())
            default = os.path.join(folder, "normalized_output.xlsx")
            self.output_path.set(default)

    # ── Logging ───────────────────────────────────────────────────────────

    def _log(self, msg: str, tag="info"):
        self.log_text.configure(state="normal")
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] {msg}\n", tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _log_clear(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    # ── Run ───────────────────────────────────────────────────────────────

    def _run(self):
        sales = self.sales_path.get().strip()
        cash = self.cash_path.get().strip()
        out = self.output_path.get().strip()

        # Validate inputs
        if not sales:
            messagebox.showerror("Missing Input", "Please select a Sales Report file.")
            return
        if not cash:
            messagebox.showerror("Missing Input", "Please select a Cash Register file.")
            return
        if not out:
            messagebox.showerror("Missing Input", "Please specify an output file path.")
            return
        if not os.path.isfile(sales):
            messagebox.showerror("File Not Found", f"Sales Report not found:\n{sales}")
            return
        if not os.path.isfile(cash):
            messagebox.showerror("File Not Found", f"Cash Register not found:\n{cash}")
            return

        self._log_clear()
        self.btn_run.configure(state="disabled")
        self.progress.start(12)
        self.status_var.set("Processing…")

        def worker():
            try:
                sales_rows, cash_rows = normalize(
                    sales, cash, out,
                    log_fn=lambda m: self.after(0, self._log, m),
                )

                def on_success():
                    self.progress.stop()
                    self.btn_run.configure(state="normal")
                    self.status_var.set(
                        f"Done — {len(sales_rows)} test rows, {len(cash_rows)} cash entries"
                    )
                    self._log(f"Output: {out}", "success")
                    messagebox.showinfo(
                        "Export Complete",
                        f"✅ Export successful!\n\n"
                        f"  • Sheet 1 (sales_report)  : {len(sales_rows)} rows\n"
                        f"  • Sheet 2 (cash_register) : {len(cash_rows)} rows\n\n"
                        f"Saved to:\n{out}",
                    )

                self.after(0, on_success)

            except Exception as exc:
                import traceback
                tb = traceback.format_exc()

                def on_error():
                    self.progress.stop()
                    self.btn_run.configure(state="normal")
                    self.status_var.set("Error — see log")
                    self._log(f"ERROR: {exc}", "error")
                    for line in tb.splitlines():
                        self._log(line, "error")
                    messagebox.showerror(
                        "Processing Error",
                        f"An error occurred:\n\n{exc}\n\nSee the log for details.",
                    )

                self.after(0, on_error)

        threading.Thread(target=worker, daemon=True).start()


# ─────────────────────────── entry point ────────────────────────────────────

if __name__ == "__main__":
    app = OPDNormalizerApp()
    app.mainloop()
